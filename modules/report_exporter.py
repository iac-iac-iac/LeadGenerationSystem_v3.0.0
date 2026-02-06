import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path
from datetime import datetime


class ReportExporter:
    """Экспорт отчётов в Excel"""

    @staticmethod
    def export_to_excel(metrics, chart_paths, output_path):
        """
        Экспорт отчёта в Excel с диаграммами

        Args:
            metrics: Словарь с метриками из Analytics
            chart_paths: Словарь с путями к диаграммам {'pie': path, 'bar': path}
            output_path: Путь для сохранения Excel
        """
        # Создание книги Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчёт"

        # Стили
        header_font = Font(size=14, bold=True)
        subheader_font = Font(size=12, bold=True)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Заголовок
        ws['A1'] = "ОТЧЁТ ПО ЛИДОГЕНЕРАЦИИ"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Дата создания: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:D2')

        # Общая статистика
        row = 4
        ws[f'A{row}'] = "1. ОБЩАЯ СТАТИСТИКА"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
        ws.merge_cells(f'A{row}:D{row}')

        row += 1
        stats = [
            ('Всего загружено', metrics.get('total_leads', 0), 'лидов'),
            ('В работе (DEAL)', metrics.get('total_deals', 0), 'сделок'),
            ('Отказы (LEAD)', metrics.get('total_rejections', 0), 'лидов'),
            ('Успешные продажи', metrics.get('successful_deals', 0), 'сделок'),
            ('Конверсия', f"{metrics.get('conversion', 0)}%", ''),
        ]

        for label, value, unit in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'C{row}'] = unit
            row += 1

        # Причины отказа
        row += 2
        ws[f'A{row}'] = "2. ПРИЧИНЫ ОТКАЗА"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
        ws.merge_cells(f'A{row}:D{row}')

        row += 1
        rejection_reasons = metrics.get('rejection_reasons', {})
        total_rejections = metrics.get('total_rejections', 1)

        if rejection_reasons:
            ws[f'A{row}'] = "Причина"
            ws[f'B{row}'] = "Количество"
            ws[f'C{row}'] = "Процент"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
            row += 1

            for reason, count in rejection_reasons.items():
                percentage = (count / total_rejections) * 100
                ws[f'A{row}'] = reason
                ws[f'B{row}'] = count
                ws[f'C{row}'] = f"{percentage:.1f}%"
                row += 1
        else:
            ws[f'A{row}'] = "Нет данных"
            row += 1

        # Топ-менеджеры
        row += 2
        ws[f'A{row}'] = "3. ТОП-МЕНЕДЖЕРЫ"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
        ws.merge_cells(f'A{row}:D{row}')

        row += 1
        top_managers = metrics.get('top_managers', {})

        if top_managers:
            ws[f'A{row}'] = "Место"
            ws[f'B{row}'] = "Менеджер"
            ws[f'C{row}'] = "Сделок"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
            row += 1

            for idx, (manager, count) in enumerate(top_managers.items(), 1):
                ws[f'A{row}'] = idx
                ws[f'B{row}'] = manager
                ws[f'C{row}'] = count
                row += 1
        else:
            ws[f'A{row}'] = "Нет данных"
            row += 1

        # Вставка диаграмм
        row += 2

        if chart_paths.get('pie') and Path(chart_paths['pie']).exists():
            ws[f'A{row}'] = "ДИАГРАММА: Причины отказа"
            ws[f'A{row}'].font = subheader_font
            row += 1

            img = XLImage(chart_paths['pie'])
            img.width = 600
            img.height = 360
            ws.add_image(img, f'A{row}')
            row += 20

        if chart_paths.get('bar') and Path(chart_paths['bar']).exists():
            ws[f'A{row}'] = "ДИАГРАММА: Топ-менеджеры"
            ws[f'A{row}'].font = subheader_font
            row += 1

            img = XLImage(chart_paths['bar'])
            img.width = 600
            img.height = 360
            ws.add_image(img, f'A{row}')

        # Установка ширины колонок
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        # Сохранение
        wb.save(output_path)
        print(f"✅ Excel отчёт сохранён: {output_path}")

        return output_path
